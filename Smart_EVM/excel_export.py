"""
excel_export.py — Export vote results to an XLSX file using openpyxl.
"""

import os
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

from config import EXPORT_PATH, DEFAULT_CANDIDATES
from database import get_vote_totals, get_all_events
from logger import get_logger

log = get_logger("smart_evm.export")


def _header_style(cell, bg: str = "1F2937", fg: str = "FFFFFF") -> None:
    cell.font = Font(bold=True, color=fg, size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def export(candidate_map: dict[int, str], out_path: str = EXPORT_PATH) -> str:
    """
    Build and save the Excel report.
    Returns the absolute path of the saved file.
    """
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    totals = get_vote_totals()
    events = get_all_events()

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    #  Sheet 1 — Summary                                                  #
    # ------------------------------------------------------------------ #
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.column_dimensions["A"].width = 6
    ws_sum.column_dimensions["B"].width = 22
    ws_sum.column_dimensions["C"].width = 14

    title_cell = ws_sum["A1"]
    title_cell.value = "SMART EVM — Vote Summary"
    title_cell.font = Font(bold=True, size=14, color="1A56DB")
    ws_sum.merge_cells("A1:C1")
    title_cell.alignment = Alignment(horizontal="center")

    ws_sum["A2"] = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_sum["A2"].font = Font(italic=True, color="6B7280", size=9)
    ws_sum.merge_cells("A2:C2")

    headers = ["ID", "Candidate", "Votes"]
    for col, h in enumerate(headers, start=1):
        _header_style(ws_sum.cell(row=4, column=col, value=h))

    total_votes = 0
    for row_idx, (cid, name) in enumerate(candidate_map.items(), start=5):
        votes = totals.get(cid, 0)
        total_votes += votes
        ws_sum.cell(row=row_idx, column=1, value=cid)
        ws_sum.cell(row=row_idx, column=2, value=name)
        ws_sum.cell(row=row_idx, column=3, value=votes)

    footer_row = 5 + len(candidate_map)
    ws_sum.cell(row=footer_row, column=2, value="TOTAL").font = Font(bold=True)
    total_cell = ws_sum.cell(row=footer_row, column=3, value=total_votes)
    total_cell.font = Font(bold=True)

    # ------------------------------------------------------------------ #
    #  Bar Chart                                                           #
    # ------------------------------------------------------------------ #
    bar = BarChart()
    bar.type = "col"
    bar.title = "Votes per Candidate"
    bar.y_axis.title = "Votes"
    bar.x_axis.title = "Candidate"
    bar.style = 10
    bar.width = 18
    bar.height = 12

    data_ref = Reference(ws_sum, min_col=3, min_row=4, max_row=4 + len(candidate_map))
    cats_ref = Reference(ws_sum, min_col=2, min_row=5, max_row=4 + len(candidate_map))
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    ws_sum.add_chart(bar, "E4")

    # ------------------------------------------------------------------ #
    #  Pie Chart                                                           #
    # ------------------------------------------------------------------ #
    pie = PieChart()
    pie.title = "Vote Distribution"
    pie.style = 10
    pie.width = 14
    pie.height = 12
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(cats_ref)
    ws_sum.add_chart(pie, "E22")

    # ------------------------------------------------------------------ #
    #  Sheet 2 — Event Log                                                 #
    # ------------------------------------------------------------------ #
    ws_log = wb.create_sheet("Event Log")
    ws_log.column_dimensions["A"].width = 6
    ws_log.column_dimensions["B"].width = 22
    ws_log.column_dimensions["C"].width = 10
    ws_log.column_dimensions["D"].width = 22
    ws_log.column_dimensions["E"].width = 14

    log_headers = ["ID", "Timestamp", "Candidate ID", "Candidate Name", "Event Type"]
    for col, h in enumerate(log_headers, start=1):
        _header_style(ws_log.cell(row=1, column=col, value=h))

    for r_idx, row in enumerate(reversed(events), start=2):
        ws_log.cell(row=r_idx, column=1, value=row["id"])
        ws_log.cell(row=r_idx, column=2, value=row["timestamp"])
        ws_log.cell(row=r_idx, column=3, value=row["candidate_id"])
        ws_log.cell(row=r_idx, column=4, value=row["candidate_name"])
        ws_log.cell(row=r_idx, column=5, value=row["event_type"])

    wb.save(out_path)
    abs_path = os.path.abspath(out_path)
    log.info("Excel exported → %s", abs_path)
    return abs_path
